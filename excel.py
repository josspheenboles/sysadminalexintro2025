import openpyxl
import sys
import os

#define read
def read(excelpath,cellnum,rownum):
    try:
        # excelpath=input('enter file path')
        #load
        wb=openpyxl.load_workbook(excelpath)
        #select sheet
        sheet=wb.active
        #read
        sheet.cell(rownum,cellnum)
    except FileNotFoundError:
        print('hjhjkhjk')
    except:
        print(sys.exc_info()[1])

#define write
def checkandcreatfile(excepath):
    if (not os.path.exists(excelpath)):
        wb =openpyxl.Workbook()
        wb.save()
def write(excelpath,sheetname,row,col,value):
    checkandcreatfile()
    wb=openpyxl.load_workbook(excelpath)        
    sheet=wb[sheetname]
    sheet.cell(row=row,column=col).value=value    
    wb.save()
